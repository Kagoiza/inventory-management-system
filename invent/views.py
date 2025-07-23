from django.db.models import Q
from django.core.mail import send_mail
from django.shortcuts import render, redirect, get_object_or_404
# Removed UserCreationForm, assuming CustomCreationForm is used
from django.contrib.auth.forms import AuthenticationForm
from .forms import CustomCreationForm
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db.models import Sum, F, Count, Q
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
import json
from django.utils.safestring import mark_safe

# Correct Model and Form Imports
from .models import ItemRequest, InventoryItem, StockTransaction
from .forms import ItemRequestForm
from .forms import InventoryItemForm
from .forms import IssueItemForm
from .forms import AdjustStockForm
# Import the new forms for return logic
from .forms import ReturnItemForm, SelectRequestForReturnForm  # NEW

import openpyxl
from openpyxl import Workbook
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.contrib.auth.models import Group
from django.contrib import messages
from django.core.paginator import Paginator


def register(request):
    if request.method == 'POST':
        # Ensure you use CustomCreationForm if that's what you intend
        form = CustomCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(
                request, "Registration successful. Please log in.")
            return redirect('login')
    else:
        form = CustomCreationForm()
    return render(request, 'invent/register.html', {'form': form})


def custom_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            if user.is_staff:
                return redirect('store_clerk_dashboard')
            else:
                return redirect('requestor_dashboard')

        else:
            messages.error(request, "Invalid username or password.")
    else:
        form = AuthenticationForm()
    return render(request, 'invent/login.html', {'form': form})


def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('login')


@login_required
def requestor_dashboard(request):
    user_requests = ItemRequest.objects.filter(
        requestor=request.user).order_by('-date_requested')

    # Fetch available inventory items for the requestor
    # Only show items with a quantity remaining > 0
    available_inventory = [item for item in InventoryItem.objects.all().order_by(
        'name') if item.quantity_remaining() > 0]

    total_requests = user_requests.count()
    approved_count = user_requests.filter(status='Approved').count()
    pending_count = user_requests.filter(status='Pending').count()
    # ADDED: Count issued and returned requests for requestor dashboard
    issued_count = user_requests.filter(status='Issued').count()
    fully_returned_count = user_requests.filter(
        status='Fully Returned').count()
    partially_returned_count = user_requests.filter(
        status='Partially Returned').count()

    return render(request, 'invent/requestor_dashboard.html', {
        'requests': user_requests,
        'total_requests': total_requests,
        'approved_count': approved_count,
        'pending_count': pending_count,
        'issued_count': issued_count,  # NEW
        'fully_returned_count': fully_returned_count,  # NEW
        'partially_returned_count': partially_returned_count,  # NEW
        # Pass available inventory (though not directly used by default dashboard content, good to have)
        'available_inventory': available_inventory,
    })


@login_required
def request_item(request):
    item_id_from_get = request.GET.get('item')

    available_inventory_queryset = InventoryItem.objects.all().order_by('name')
    available_inventory_list = [
        item for item in available_inventory_queryset if item.quantity_remaining() > 0]

    available_item_ids = [item.id for item in available_inventory_list]
    filtered_inventory_queryset = InventoryItem.objects.filter(
        id__in=available_item_ids).order_by('name')

    # JS-safe JSON data for showing item quantity/condition
    available_inventory_json = mark_safe(json.dumps([
        {
            "id": item.id,
            "quantity_remaining": item.quantity_remaining(),
            "condition": item.condition or "N/A"
        }
        for item in available_inventory_list
    ]))

    if request.method == 'POST':
        form = ItemRequestForm(request.POST)
        form.fields['item'].queryset = filtered_inventory_queryset

        if form.is_valid():
            item_request = form.save(commit=False)
            item_request.requestor = request.user
            item_request.name = item_request.item.name
            item_request.save()

            # Send email notification to the requestor
            send_mail(
                subject='Item Request Confirmation',
                message=(
                    f"Dear {request.user.first_name or request.user.username},\n\n"
                    f"Your request for item \"{item_request.item.name}\" has been successfully submitted.\n"
                    f"We will notify you once it is reviewed or issued.\n\n"
                    f"Thank you,\nInventory Management Team"
                ),
                from_email=None,  # Uses DEFAULT_FROM_EMAIL
                recipient_list=[request.user.email],
                fail_silently=False,
            )

            messages.success(request, "Item request submitted successfully!")
            return redirect('requestor_dashboard')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        initial_data = {}
        if item_id_from_get and item_id_from_get.isdigit():
            try:
                item = InventoryItem.objects.get(id=int(item_id_from_get))
                if item.id in available_item_ids:
                    initial_data['item'] = item.id
            except InventoryItem.DoesNotExist:
                pass

        form = ItemRequestForm(initial=initial_data)
        form.fields['item'].queryset = filtered_inventory_queryset

    return render(request, 'invent/request_item.html', {
        'form': form,
        'available_inventory': available_inventory_list,
        'available_inventory_json': available_inventory_json,
    })


@login_required
def cancel_request(request, request_id):
    item_request = get_object_or_404(
        ItemRequest, id=request_id, requestor=request.user)

    # Allow cancellation for Pending, Approved, and even Issued if you decide to revert stock on cancellation
    # For now, keeping your original logic to only allow Pending cancellation
    if item_request.status == 'Pending':
        if request.method == 'POST':
            item_request.status = 'Cancelled'
            # item_request.processed_by = request.user # This field doesn't exist on ItemRequest in your models.py
            # item_request.processed_at = timezone.now() # This field doesn't exist on ItemRequest in your models.py
            item_request.save()  # The save method will handle email notification for 'Cancelled' status
            messages.success(
                request, f"Request for '{item_request.item.name}' (ID: {request_id}) has been cancelled.")
            return redirect('requestor_dashboard')
        else:
            context = {
                'item_request': item_request
            }
            return render(request, 'invent/cancel_request_confirm.html', context)
    else:
        messages.error(
            request, f"Request for '{item_request.item.name}' (ID: {request_id}) cannot be cancelled because its status is '{item_request.status}'.")
        return redirect('requestor_dashboard')


# --- Store Clerk Functionality ---

@login_required
@permission_required('invent.view_inventoryitem', raise_exception=True)
def store_clerk_dashboard(request):
    inventory_items = InventoryItem.objects.all()

    total_inventory_items = inventory_items.aggregate(
        total_sum=Sum('quantity_total'))['total_sum'] or 0
    items_issued = inventory_items.aggregate(total_issued=Sum('quantity_issued'))[
        'total_issued'] or 0
    items_returned = inventory_items.aggregate(
        total_returned=Sum('quantity_returned'))['total_returned'] or 0

    items_for_dashboard = inventory_items.order_by('-created_at')[:5]

    pending_requests_count = ItemRequest.objects.filter(
        status='Pending').count()

    # Count of Issued Requests that are not fully returned yet (eligible for return)
    issued_but_not_fully_returned_count = ItemRequest.objects.filter(
        status='Issued'
    ).exclude(
        # Exclude if original quantity <= returned quantity
        quantity__lte=F('returned_quantity')
    ).count()

    context = {
        'total_items': total_inventory_items,
        'items_issued': items_issued,
        'items_returned': items_returned,
        'items': items_for_dashboard,
        'pending_requests_count': pending_requests_count,
        # NEW for dashboard
        'issued_but_not_fully_returned_count': issued_but_not_fully_returned_count,
    }
    return render(request, 'invent/store_clerk_dashboard.html', context)

# invent/views.py

@login_required
@permission_required('invent.view_inventoryitem', raise_exception=True)
def inventory_list_view(request):
    """
    Displays a list of all inventory items with search and pagination.
    """
    query = request.GET.get('q', '')
    items = InventoryItem.objects.all().order_by('name')

    if query:
        items = items.filter(
            Q(name__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(category__icontains=query) 
        )                                    

    # Add pagination
    paginator = Paginator(items, 50)  # Show 50 items per page
    page_number = request.GET.get('page')
    page_obj  = paginator.get_page(page_number) 

    context = {
        'page_obj': page_obj,  # Now passing the paginated object as 'page_obj'
        'query': query,
    }
    return render(request, 'invent/list_inventory_items.html', context)

@login_required
@permission_required('invent.view_inventoryitem', raise_exception=True)
def manage_stock(request):
    form = InventoryItemForm()

    if request.method == 'POST':
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            new_item = form.save(commit=False)
            new_item.created_by = request.user
            new_item.save()
            messages.success(
                request, f'Inventory item "{new_item.name}" added successfully.')
            return redirect('manage_stock')
        else:
            messages.error(
                request, "Error adding item. Please check the form.")

    query = request.GET.get('q', '')
    items = InventoryItem.objects.all().order_by('name')

    if query:
        items = items.filter(
            Q(name__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(category__icontains=query)
        )

    context = {
        'form': form,
        'items': items,
        'query': query,
    }
    return render(request, 'invent/manage_stock.html', context)


@login_required
@permission_required('invent.change_inventoryitem', raise_exception=True)
def edit_item(request, item_id):
    item = get_object_or_404(InventoryItem, id=item_id)
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, instance=item)
        if form.is_valid():
            if not item.created_by:
                item.created_by = request.user
            form.save()
            messages.success(
                request, f'Inventory Item "{item.name}" updated successfully!')
            return redirect('manage_stock')
        else:
            messages.error(
                request, "Error updating item. Please correct the errors below.")
    else:
        form = InventoryItemForm(instance=item)

    context = {
        'form': form,
        'item': item,
    }
    return render(request, 'invent/edit_item.html', context)


@login_required
@permission_required('invent.can_issue_item', raise_exception=True)
def issue_item(request):
    # Fetch all requests for display in the "All Requests" tab
    all_requests = ItemRequest.objects.all().order_by('-date_requested')

    # Filter for requests that are Pending or Approved to show in the primary tab
    pending_and_approved_requests = all_requests.filter(
        status__in=['Pending', 'Approved'])

    # Initialize the direct issue form (for cases not coming from a request)
    form = IssueItemForm()  # This is the direct issue form

    if request.method == 'POST':
        action = request.POST.get('action')
        request_id = request.POST.get('request_id')

        # This block handles actions (approve, reject, issue) on existing requests
        if action and request_id:
            item_request = get_object_or_404(ItemRequest, id=request_id)
            try:
                with transaction.atomic():
                    if action == 'approve':
                        if item_request.status == 'Pending':  # Ensure only pending requests can be approved
                            item_request.status = 'Approved'
                            # item_request.processed_by = request.user # Add these fields to ItemRequest model if you need them
                            # item_request.processed_at = timezone.now()
                            item_request.save()
                            messages.success(
                                request, f"Request ID {request_id} ({item_request.item.name}) approved.")
                        else:
                            messages.warning(
                                request, f"Request ID {request_id} is '{item_request.status}' and cannot be approved.")

                    elif action == 'reject':
                        # Allow rejecting from Pending or Approved states
                        if item_request.status in ['Pending', 'Approved']:
                            item_request.status = 'Rejected'
                            # item_request.processed_by = request.user
                            # item_request.processed_at = timezone.now()
                            item_request.save()
                            messages.warning(
                                request, f"Request ID {request_id} ({item_request.item.name}) rejected.")
                        else:
                            messages.warning(
                                request, f"Request ID {request_id} is '{item_request.status}' and cannot be rejected.")

                    elif action == 'issue_from_request':
                        # *** CRITICAL CHANGE HERE: Ensure status is 'Approved' to issue ***
                        if item_request.status != 'Approved':
                            messages.error(
                                request, f"Cannot issue for request ID {request_id}. It must be 'Approved'. Current status: '{item_request.status}'.")
                            return redirect('issue_item')

                        item_to_issue_obj = item_request.item

                        if item_to_issue_obj.quantity_remaining() < item_request.quantity:
                            messages.error(
                                request, f"Insufficient stock for '{item_to_issue_obj.name}'. Requested: {item_request.quantity}, Available: {item_to_issue_obj.quantity_remaining()}.")
                            return redirect('issue_item')

                        # Deduct quantity_available and update quantity_issued
                        item_to_issue_obj.quantity_issued = F(
                            'quantity_issued') + item_request.quantity
                        item_to_issue_obj.save(
                            update_fields=['quantity_issued'])

                        StockTransaction.objects.create(
                            item=item_to_issue_obj,
                            transaction_type='Issue',
                            # Store positive quantity for Issue transactions for consistency in StockTransaction
                            quantity=item_request.quantity,
                            # and reflect change in InventoryItem quantities by F() expressions.
                            item_request=item_request,  # LINK THE TRANSACTION TO THE REQUEST
                            issued_to=item_request.requestor.username,
                            reason=f"Issued for request ID: {item_request.id} ({item_request.item.name})",
                            recorded_by=request.user
                        )
                        item_request.status = 'Issued'
                        # item_request.processed_by = request.user
                        # item_request.processed_at = timezone.now()
                        item_request.save()  # This save will now set date_issued and send email
                        messages.success(
                            request, f'Request ID {item_request.id} ({item_request.item.name}) issued and marked as Issued.')
                    else:
                        messages.error(request, "Invalid request action.")

            except Exception as e:
                messages.error(request, f"Error processing request: {e}")

            return redirect('issue_item')

        # This block handles the direct issue form submission (not tied to a request)
        # Re-initialize the form with POST data for direct issue
        form = IssueItemForm(request.POST)
        if form.is_valid():
            item_to_issue = form.cleaned_data['item']
            quantity = form.cleaned_data['quantity']
            issued_to = form.cleaned_data['issued_to']

            try:
                with transaction.atomic():
                    if item_to_issue.quantity_remaining() < quantity:
                        messages.error(
                            request, f"Not enough stock for {item_to_issue.name}. Available: {item_to_issue.quantity_remaining()}.")
                        # No redirect here, so form errors can be displayed
                        # This means you need to pass the context again
                        context = {
                            'form': form,  # Pass the form with errors back
                            'all_requests': all_requests,
                            'pending_and_approved_requests': pending_and_approved_requests,
                            'approved_requests': all_requests.filter(status='Approved'),
                            'issued_requests': all_requests.filter(status='Issued'),
                            'rejected_requests': all_requests.filter(status='Rejected'),
                            'pending_requests': all_requests.filter(status='Pending'),
                        }
                        return render(request, 'invent/issue_item.html', context)

                    item_to_issue.quantity_issued = F(
                        'quantity_issued') + quantity
                    item_to_issue.save(update_fields=['quantity_issued'])

                    StockTransaction.objects.create(
                        item=item_to_issue,
                        transaction_type='Issue',
                        quantity=quantity,  # Store positive quantity for Issue transactions
                        issued_to=issued_to,
                        reason=f"Direct issue to {issued_to}. ",
                        recorded_by=request.user
                    )
                messages.success(
                    request, f'{quantity} x {item_to_issue.name} successfully issued to {issued_to}.')
                return redirect('issue_item')
            except Exception as e:
                messages.error(request, f"Error issuing item: {e}")
        else:
            messages.error(
                request, "Please correct the errors in the direct issue form.")
            # Important: If the form is invalid, you must render the template
            # and pass the form back so its errors can be displayed.
            context = {
                'form': form,  # Pass the form with errors back
                'all_requests': all_requests,
                'pending_and_approved_requests': pending_and_approved_requests,
                'approved_requests': all_requests.filter(status='Approved'),
                'issued_requests': all_requests.filter(status='Issued'),
                'rejected_requests': all_requests.filter(status='Rejected'),
                'pending_requests': all_requests.filter(status='Pending'),
            }
            return render(request, 'invent/issue_item.html', context)

    # GET request: Render the page with the initial data
    context = {
        'form': form,  # Ensure the form is passed for GET requests as well
        'all_requests': all_requests,
        'pending_and_approved_requests': pending_and_approved_requests,
        'approved_requests': all_requests.filter(status='Approved'),
        'issued_requests': all_requests.filter(status='Issued'),
        'rejected_requests': all_requests.filter(status='Rejected'),
        'pending_requests': all_requests.filter(status='Pending'),
    }
    return render(request, 'invent/issue_item.html', context)


@login_required
def request_summary(request):
    # Filter all requests to only those made by the logged-in requestor
    user_requests = ItemRequest.objects.filter(requestor=request.user)

    total_requests = user_requests.count()
    pending_requests = user_requests.filter(status='Pending').count()
    approved_requests = user_requests.filter(status='Approved').count()
    issued_requests = user_requests.filter(status='Issued').count()
    rejected_requests = user_requests.filter(status='Rejected').count()
    # NEW: Add counts for returned statuses
    partially_returned_requests = user_requests.filter(
        status='Partially Returned').count()
    fully_returned_requests = user_requests.filter(
        status='Fully Returned').count()
    # Sum of all returned quantities across all requests by this user
    total_returned_quantity_by_user = user_requests.aggregate(
        Sum('returned_quantity'))['returned_quantity__sum'] or 0

    requests_by_status = user_requests.values(
        'status').annotate(count=Count('id')).order_by('status')

    requests_by_item = user_requests.values('item__name').annotate(
        count=Sum('quantity')).order_by('-count')[:10]

    # No need to show top requestors to the current requestor (omit or just show current user’s total)
    # Alternatively, show how many times *they* requested:
    requests_by_requestor = user_requests.values(
        'requestor__username').annotate(count=Count('id'))

    context = {
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'approved_requests': approved_requests,
        'issued_requests': issued_requests,
        'rejected_requests': rejected_requests,
        'partially_returned_requests': partially_returned_requests,  # NEW
        'fully_returned_requests': fully_returned_requests,       # NEW
        'total_returned_quantity_by_user': total_returned_quantity_by_user,  # NEW
        'requests_by_status': requests_by_status,
        'requests_by_item': requests_by_item,
        'requests_by_requestor': requests_by_requestor,
    }
    return render(request, 'invent/request_summary.html', context)


@login_required
@permission_required('invent.change_inventoryitem', raise_exception=True)
def adjust_stock(request):
    if request.method == 'POST':
        form = AdjustStockForm(request.POST)
        if form.is_valid():
            item = form.cleaned_data['item']
            adjustment_quantity = form.cleaned_data['adjustment_quantity']
            reason = form.cleaned_data['reason']

            try:
                with transaction.atomic():
                    item.quantity_total = F(
                        'quantity_total') + adjustment_quantity
                    item.save(update_fields=['quantity_total'])

                    # Adjusted transaction_type for clarity. 'Adjustment' is better.
                    # quantity will be positive for adding, negative for removing.
                    transaction_type = 'Adjustment'

                    StockTransaction.objects.create(
                        item=item,
                        transaction_type=transaction_type,
                        quantity=adjustment_quantity,  # Store actual adjustment value
                        reason=reason,
                        recorded_by=request.user
                    )
                messages.success(
                    request, f'Stock for {item.name} adjusted by {adjustment_quantity}. New total: {item.quantity_total}.')
                return redirect('adjust_stock')
            except Exception as e:
                messages.error(request, f"Error adjusting stock: {e}")
        else:
            messages.error(
                request, "Please correct the errors in the adjustment form.")
    else:
        form = AdjustStockForm()

    # Filter for 'Adjustment' transactions for this display, as 'Issue' and 'Return' will be handled elsewhere
    recent_transactions = StockTransaction.objects.filter(
        transaction_type='Adjustment').order_by('-transaction_date')[:10]

    context = {
        'form': form,
        'recent_transactions': recent_transactions,
    }
    return render(request, 'invent/adjust_stock.html', context)


@login_required
@permission_required('invent.change_inventoryitem', raise_exception=True)
def reports(request):
    # This view seems to be a placeholder or generic reports view.
    # The actual data is fetched in `reports_view` below.
    # It might be better to consolidate, or have this render a template that
    # includes components using the data from `reports_view`.
    # For now, I'll update reports_view and keep this for URL mapping if needed.
    return render(request, 'invent/reports.html')


# Consolidated reports_view for cleaner logic and direct use.
# Added permission requirement for store clerks.
@login_required
# Assuming clerks need to see reports
@permission_required('invent.view_inventoryitem', raise_exception=True)
def reports_view(request):
    context = {
        'total_items': InventoryItem.objects.aggregate(total=Sum('quantity_total'))['total'] or 0,
        'total_requests': ItemRequest.objects.count(),
        'pending_count': ItemRequest.objects.filter(status='Pending').count(),
        'approved_count': ItemRequest.objects.filter(status='Approved').count(),
        'issued_count': ItemRequest.objects.filter(status='Issued').count(),
        'rejected_count': ItemRequest.objects.filter(status='Rejected').count(),
        # MODIFIED: Calculate returned_count from ItemRequest statues
        'fully_returned_count': ItemRequest.objects.filter(status='Fully Returned').count(),
        'partially_returned_count': ItemRequest.objects.filter(status='Partially Returned').count(),
        # Sum of actual quantities returned via transactions or the ItemRequest.returned_quantity field
        'total_returned_quantity_all_items': ItemRequest.objects.aggregate(total_returned=Sum('returned_quantity'))['total_returned'] or 0,


        # Top 2 requested items
        'top_requested_items': (
            ItemRequest.objects.values('item__name')
            .annotate(request_count=Count('id'))
            .order_by('-request_count')[:2]
        )
    }
    return render(request, 'invent/reports.html', context)


@login_required
@permission_required('invent.add_inventoryitem', raise_exception=True)
def upload_inventory(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # File extension check
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect('upload_inventory')

        # Save file temporarily
        file_name = default_storage.save(
            excel_file.name, ContentFile(excel_file.read()))
        file_path = default_storage.path(file_name)

        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            success_count = 0
            skipped_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                # Ensure the number of columns matches your excel structure
                # This needs to be consistent with the actual columns in your Excel file.
                # Assuming the order is: name, serial, category, condition, status, total, issued, returned
                name, serial, category, condition, status, total, issued, returned = row

                # Skip if required fields are missing
                if not name:
                    skipped_count += 1
                    continue

                try:
                    InventoryItem.objects.create(
                        name=name.strip(),
                        serial_number=(serial or "").strip(),
                        category=(category or "").strip(),
                        condition=condition or "Serviceable",
                        status=status or "In Stock",
                        quantity_total=int(total or 0),
                        quantity_issued=int(issued or 0),
                        # Ensure this maps correctly
                        quantity_returned=int(returned or 0),
                        created_by=request.user
                    )
                    success_count += 1
                except Exception as e:
                    messages.error(
                        request, f"Error processing row for item '{name}': {e}")
                    skipped_count += 1  # Skip row if conversion or save fails

            # Feedback message
            messages.success(
                request,
                f"{success_count} item(s) uploaded successfully. {skipped_count} row(s) skipped due to errors or missing data."
            )
            return redirect('upload_inventory')

        except Exception as e:
            messages.error(request, f"Failed to process Excel file: {e}")
            return redirect('upload_inventory')
        finally:
            # Clean up the temporary file
            if default_storage.exists(file_name):
                default_storage.delete(file_name)

    return render(request, 'invent/upload_inventory.html')

# <-- Reports Section ---


def total_requests(request):
    status_filter = request.GET.get('status')
    # Filter ItemRequest by relevant statuses for the "All Requests" report
    # and exclude partially/fully returned if you only want truly active ones.
    # For a total request list, usually all are included unless specified.
    requests = ItemRequest.objects.all()

    if status_filter:
        requests = requests.filter(status=status_filter)

    paginator = Paginator(requests.order_by('-date_requested'), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
    }
    return render(request, 'invent/total_requests.html', context)

# Export


def export_total_requests(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    status_filter = request.GET.get('status')
    queryset = ItemRequest.objects.all()

    if status_filter:
        queryset = queryset.filter(status=status_filter)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Item Requests"

    # Define headers
    headers = ['Requested By', 'Item', 'Quantity Requested',
               'Quantity Returned', 'Date Requested', 'Status']  # MODIFIED headers
    ws.append(headers)

    # Add data rows
    for item_req in queryset:  # Changed 'item' to 'item_req' for clarity
        ws.append([
            item_req.requestor.username,
            item_req.item.name,
            item_req.quantity,
            item_req.returned_quantity,  # NEW column
            item_req.date_requested.strftime('%Y-%m-%d'),
            item_req.status
        ])

    # Adjust column widths (optional)
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # Set up HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=total_requests.xlsx'
    wb.save(response)
    return response


def export_inventory_items(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=inventory_items.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Items"

    # Header row
    ws.append([
        'Item ID',
        'Item Name',
        'Serial Number',
        'Category',
        'Condition',
        'Status',
        'Total Qty',
        'Issued Qty',
        # Clarified this is the aggregate field on InventoryItem
        'Returned Qty (Aggregate)',
        'Available Qty (Calculated)',  # Clarified this is calculated
    ])

    # Data rows
    for item in InventoryItem.objects.all():
        ws.append([
            item.id,
            item.name,
            item.serial_number if item.serial_number else '',  # Corrected access
            item.category if item.category else 'N/A',
            item.condition if item.condition else 'N/A',
            # This status might be 'In Stock', 'Low Stock', etc.
            item.status if item.status else 'N/A',
            item.quantity_total,
            item.quantity_issued,
            item.quantity_returned,
            item.quantity_remaining(),  # Using your existing method which is (total - issued)
            # If you want it to be (total - issued + returned_to_total), adjust quantity_total logic
        ])

    wb.save(response)
    return response


# --- NEW RETURN LOGIC VIEWS ---

@login_required
# Assuming clerks handle returns
@permission_required('invent.can_issue_item', raise_exception=True)
def list_issued_requests_for_return(request):
    # Filter ItemRequests that have been 'Issued' and where the returned_quantity
    # is less than the original requested quantity (meaning not fully returned yet)
    issued_requests = ItemRequest.objects.filter(
        status='Issued'
    ).exclude(
        # Exclude if original quantity <= returned quantity
        quantity__lte=F('returned_quantity')
    ).select_related('item', 'requestor').order_by('-date_issued')

    context = {
        'issued_requests': issued_requests,
        'title': 'Issued Items for Return'
    }
    return render(request, 'invent/list_issued_requests_for_return.html', context)


@login_required
# Assuming clerks handle returns
@permission_required('invent.can_issue_item', raise_exception=True)
def process_return_for_request(request, request_id):
    # Retrieve the ItemRequest, ensuring it's an 'Issued' request and not fully returned
    item_request = get_object_or_404(
        ItemRequest.objects.filter(status='Issued').exclude(
            quantity__lte=F('returned_quantity')),
        id=request_id
    )

    if request.method == 'POST':
        form = ReturnItemForm(request.POST, item_request=item_request)
        if form.is_valid():
            returned_quantity = form.cleaned_data['returned_quantity']
            # Use .get for optional fields
            reason = form.cleaned_data.get('reason')

            # This check is also in the form's clean method, but a double-check here is fine.
            if returned_quantity > item_request.quantity_to_be_returned():
                messages.error(
                    request, "Cannot return more than the remaining issued quantity for this request.")
                return render(request, 'invent/process_return_for_request.html', {'form': form, 'item_request': item_request})

            try:
                with transaction.atomic():
                    # 1. Update InventoryItem's quantity_issued and quantity_total
                    # When an item is returned, it should be deducted from 'quantity_issued'
                    # and added back to 'quantity_total' (available stock).
                    item = item_request.item
                    item.quantity_issued = F(
                        'quantity_issued') - returned_quantity  # Reduce issued count
                    # Increase total available stock
                    item.quantity_total = F(
                        'quantity_total') + returned_quantity
                    # Update aggregate returned count on InventoryItem
                    item.quantity_returned = F(
                        'quantity_returned') + returned_quantity
                    item.save(update_fields=[
                              'quantity_issued', 'quantity_total', 'quantity_returned'])
                    # Reload the item instance to get the updated values after F() expression save
                    item.refresh_from_db()

                    # 2. Create a StockTransaction for the return
                    StockTransaction.objects.create(
                        item=item,
                        transaction_type='Return',
                        quantity=returned_quantity,  # Store the positive quantity that was returned
                        item_request=item_request,  # Link to the original request
                        reason=reason,
                        recorded_by=request.user
                    )

                    # 3. Update ItemRequest's returned_quantity and status
                    item_request.returned_quantity = F(
                        'returned_quantity') + returned_quantity
                    # Save before checking status to ensure F() is applied
                    item_request.save(update_fields=['returned_quantity'])
                    # Refresh to get the actual updated returned_quantity
                    item_request.refresh_from_db()

                    if item_request.returned_quantity == item_request.quantity:
                        item_request.status = 'Fully Returned'
                    elif item_request.returned_quantity > 0:  # Check > 0 after the update
                        item_request.status = 'Partially Returned'
                    item_request.save()  # This save will trigger the email notification for status change

                    messages.success(
                        request, f"Successfully returned {returned_quantity} of {item_request.item.name} for request ID {item_request.id}.")
                    # Redirect back to the list
                    return redirect('list_issued_requests_for_return')

            except Exception as e:
                messages.error(
                    request, f"An error occurred while processing the return: {e}")
    else:
        # Initialize form with maximum allowed return quantity
        form = ReturnItemForm(item_request=item_request)

    context = {
        'form': form,
        'item_request': item_request,
        'title': f'Return Item for Request {item_request.id}'
    }
    return render(request, 'invent/process_return_for_request.html', context)
